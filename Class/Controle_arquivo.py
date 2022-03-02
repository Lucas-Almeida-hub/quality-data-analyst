from openpyxl import load_workbook


Obj_arquivo=""
arquivos=""

class File_read:

    def __init__(self) -> None:
        global Obj_arquivo
        arquivos = load_workbook('/home/lucas/Helem/BD/Review.xlsx') 
        Obj_arquivo = arquivos.active

    def read(self,row):
        global Obj_arquivo
        Result = Obj_arquivo.cell(row,column=1)
        print(Result.value)
        return(Result.value)




